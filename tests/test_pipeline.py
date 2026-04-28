"""Unit tests for the PH Hunter Pipeline."""
import json
import sys
import os
import tempfile
from pathlib import Path
from datetime import datetime, timezone, timedelta

sys.path.insert(0, str(Path(__file__).parent.parent))

# Use dates relative to today to avoid stale test data
_NOW = datetime.now(timezone.utc)
_RECENT = (_NOW - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S+00:00")   # 30 days ago — pass
_OLD    = (_NOW - timedelta(days=200)).strftime("%Y-%m-%dT%H:%M:%S+00:00")  # 200 days ago — fail


def test_config_collector():
    from src.experts.config_collector import collect_config
    cfg = collect_config(["AI", "SaaS"], min_followers=500, search_mode="keywords")
    assert cfg["min_followers"] == 500
    assert "AI" in cfg["topics"]
    assert cfg["search_mode"] == "keywords"
    print("  [OK] config_collector: valid params")

    try:
        collect_config(["AI"], priority_threshold=60, secondary_threshold=80)
        assert False, "Should have raised"
    except AssertionError:
        pass
    print("  [OK] config_collector: rejects priority < secondary")


def test_filter_hunters_gates():
    from src.experts.hunter_filter import filter_hunters
    hunters = [
        {"username": "a", "ph_followers": 5000, "products_hunted": 20, "last_activity_date": _RECENT},
        {"username": "b", "ph_followers": 100,  "products_hunted": 20, "last_activity_date": _RECENT},   # low followers
        {"username": "c", "ph_followers": 5000, "products_hunted": 2,  "last_activity_date": _RECENT},   # few hunts
        {"username": "d", "ph_followers": 5000, "products_hunted": 20, "last_activity_date": _OLD},      # inactive
    ]
    passed, killed = filter_hunters(hunters, min_followers=1000, min_hunts=10, activity_days=90)
    assert len(passed) == 1 and passed[0]["username"] == "a",         f"Expected only 'a', got {[h['username'] for h in passed]}"
    print(f"  [OK] filter: only 'a' passed (recent={_RECENT[:10]}, old={_OLD[:10]})")
    assert killed["followers"] == 1
    print("  [OK] filter: followers gate kills correctly")
    assert killed["hunts"] == 1
    print("  [OK] filter: hunts gate kills correctly")
    assert killed["activity"] == 1
    print("  [OK] filter: activity gate HARD KILLS correctly")

    # Test missing date does NOT kill
    h_no_date = [{"username": "x", "ph_followers": 5000, "products_hunted": 20, "last_activity_date": None}]
    passed2, _ = filter_hunters(h_no_date, 1000, 10, 90)
    assert len(passed2) == 1
    print("  [OK] filter: missing activity_date does not kill")


def test_scorer():
    from src.experts.hunter_scorer import score_hunters
    hunters = [
        {"username": "top", "ph_followers": 25000, "products_hunted": 60,
         "categories": ["AI", "SaaS"], "avg_upvotes_last_10": 1200,
         "twitter_username": "topuser", "x_followers": 0},
        {"username": "low", "ph_followers": 1200, "products_hunted": 12,
         "categories": ["cooking"], "avg_upvotes_last_10": 15,
         "twitter_username": "", "x_followers": 0},
    ]
    scored = score_hunters(hunters, ["AI", "SaaS"], 80, 60)
    scores = {h["username"]: h["score"] for h in scored}
    assert scores["top"] > scores["low"], f"top={scores['top']} should > low={scores['low']}"
    print(f"  [OK] scorer: top ({scores['top']}) > low ({scores['low']})")
    top_h = next(h for h in scored if h["username"] == "top")
    assert "segment" in top_h
    print(f"  [OK] scorer: segment assigned ({top_h['segment']})")
    assert 0 <= scores["top"] <= 100 and 0 <= scores["low"] <= 100
    print("  [OK] scorer: scores in valid 0-100 range")

    # Zero category fit
    h_nocat = [{"username": "z", "ph_followers": 10000, "products_hunted": 20,
                "categories": [], "avg_upvotes_last_10": 100,
                "twitter_username": "", "x_followers": 0}]
    scored_z = score_hunters(h_nocat, ["AI"], 80, 60)
    assert scored_z[0]["score_breakdown"]["category_fit_pct"] == 0.0
    print("  [OK] scorer: zero category fit when no categories")


def test_excel_exporter():
    import openpyxl
    from src.experts.excel_exporter import export_to_excel
    test_hunters = [
        {"username": "alice", "name": "Alice Smith", "ph_url": "https://producthunt.com/@alice",
         "linkedin_url": "https://linkedin.com/in/alice", "score": 85, "segment": "Priority",
         "ph_followers": 12000, "x_followers": 8000, "categories": ["AI", "SaaS"],
         "is_eu": True, "country": "Germany", "last_activity_date": "2026-03-01",
         "products_hunted": 35, "avg_upvotes_last_10": 700, "twitter_username": "alice_tw"},
        {"username": "bob", "name": "Bob Jones", "ph_url": "https://producthunt.com/@bob",
         "linkedin_url": None, "score": 72, "segment": "Secondary",
         "ph_followers": 3000, "x_followers": 500, "categories": ["tools"],
         "is_eu": False, "country": "US", "last_activity_date": None,
         "products_hunted": 18, "avg_upvotes_last_10": 120, "twitter_username": "bobjones"},
        {"username": "carol", "name": "Carol Lee", "ph_url": "https://producthunt.com/@carol",
         "linkedin_url": "https://linkedin.com/in/carol", "score": 65, "segment": "Secondary",
         "ph_followers": 2500, "x_followers": 200, "categories": ["productivity"],
         "is_eu": False, "country": "UK", "last_activity_date": "2026-02-15",
         "products_hunted": 22, "avg_upvotes_last_10": 300, "twitter_username": "carollee"},
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name
    try:
        out_path = export_to_excel(test_hunters, output_path=tmp_path)
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        assert ws.max_row == 3, f"Expected header + 2 data rows (alice+carol), got {ws.max_row}"
        print(f"  [OK] excel: exports only hunters with LinkedIn ({ws.max_row - 1} rows)")
        row2_name = ws.cell(row=2, column=1).value
        assert row2_name == "Alice Smith", f"Expected Alice first (score 85), got {row2_name}"
        print(f"  [OK] excel: sorted by score DESC (first: {row2_name})")
        assert "Summary" in wb.sheetnames
        print("  [OK] excel: Summary sheet created")
        linkedin_cell = ws.cell(row=2, column=3)
        assert linkedin_cell.hyperlink is not None
        print("  [OK] excel: LinkedIn URL is hyperlink")
    finally:
        os.unlink(tmp_path)


if __name__ == "__main__":
    print("\n" + "="*50)
    print("🧪 PH Hunter Pipeline — Unit Tests")
    print("="*50)
    ok = err = 0
    for fn in [test_config_collector, test_filter_hunters_gates, test_scorer, test_excel_exporter]:
        print(f"\n▶ {fn.__name__}")
        try:
            fn(); ok += 1
        except Exception as e:
            print(f"  [FAIL] {e}"); err += 1
    print(f"\n{'✅ ALL PASSED' if err==0 else f'❌ {err} FAILED'} | {ok} passed, {err} failed")
    sys.exit(0 if err == 0 else 1)
