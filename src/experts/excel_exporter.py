"""Excel exporter — writes scored hunters to .xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

def export_to_excel(hunters, output_path="", include_below=False):
    rows = [h for h in hunters if h.get("linkedin_url")]
    if not include_below:
        rows = [h for h in rows if h.get("segment") in ("Priority", "Secondary")]
    rows.sort(key=lambda h: h.get("score", 0), reverse=True)
    if not output_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = str(Path.home() / "Downloads" / ("ph_hunters_" + ts + ".xlsx"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PH Hunters"
    HDRS = ["Name", "PH Profile", "LinkedIn", "Score", "Segment",
            "PH Followers", "X Followers", "Categories", "Geo",
            "Last Activity", "Products Hunted", "Avg Upvotes", "Twitter"]
    WIDTHS = [22, 38, 45, 8, 12, 14, 12, 42, 22, 14, 16, 16, 18]
    hf = Font(bold=True, color="FFFFFF", name="Calibri")
    hfill = PatternFill(start_color="1E2A3A", end_color="1E2A3A", fill_type="solid")
    for ci, (h, w) in enumerate(zip(HDRS, WIDTHS), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    pri_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sec_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    url_font = Font(color="0563C1", underline="single", name="Calibri")
    for ri, h in enumerate(rows, 2):
        seg = h.get("segment", "")
        fill = pri_fill if seg == "Priority" else (sec_fill if seg == "Secondary" else None)
        country = h.get("country", "")
        geo = ("EU " + country) if h.get("is_eu") else (country or "N/A")
        cats = h.get("categories") or []
        cats_str = ", ".join(str(c) for c in cats[:5])
        avg_up = h.get("avg_upvotes_last_10", 0)
        data = [
            h.get("name", h.get("username", "")),
            h.get("ph_url", ""),
            h.get("linkedin_url", ""),
            h.get("score", 0),
            seg,
            h.get("ph_followers", 0),
            h.get("x_followers", 0),
            cats_str,
            geo,
            str(h.get("last_activity_date", ""))[:10],
            h.get("products_hunted", 0),
            round(avg_up, 1) if avg_up else 0,
            h.get("twitter_username", "")
        ]
        for ci, v in enumerate(data, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            if fill:
                c.fill = fill
            c.alignment = Alignment(vertical="center")
            if ci in (2, 3) and v and str(v).startswith("http"):
                c.hyperlink = str(v)
                c.font = url_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ("A1:" + get_column_letter(len(HDRS)) +
                          str(max(len(rows), 1) + 1))
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "PH Hunter Pipeline Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    pri_c = sum(1 for h in rows if h.get("segment") == "Priority")
    sec_c = sum(1 for h in rows if h.get("segment") == "Secondary")
    eu_c = sum(1 for h in rows if h.get("is_eu"))
    no_li = sum(1 for h in hunters if not h.get("linkedin_url"))
    summary = [
        ("", ""),
        ("Total exported (with LinkedIn)", len(rows)),
        ("Priority segment", pri_c),
        ("Secondary segment", sec_c),
        ("EU hunters", eu_c),
        ("Excluded (no LinkedIn)", no_li),
        ("", ""),
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for ri2, (lbl, val) in enumerate(summary, 2):
        ws2.cell(row=ri2, column=1, value=lbl)
        c = ws2.cell(row=ri2, column=2, value=val)
        if lbl:
            c.font = Font(bold=True, name="Calibri")
    wb.save(output_path)
    return output_path
