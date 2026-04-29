"""
Self-contained scoring + Excel export.
Reads enricher checkpoint, scores inline, builds Excel inline.
No nested HTTP calls — works around Extella nested expert timeout.

Key design decision:
- avg_upvotes always 0 (can't collect without visiting product pages)
- Max achievable score = 68-79 (avg_upvotes=20% wasted)
- Recommended thresholds: priority=50, secondary=25

Scoring weights:
  PH followers   25% (log-scale 1K-50K)
  Category fit   25% (partial match with target categories)
  Avg upvotes    20% (ALWAYS 0 — limitation)
  Products hunted 15% (linear 10-100)
  X followers    10% (log-scale 100-100K)
"""
import json
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


EU_CODES = {"at","be","bg","hr","cy","cz","dk","ee","fi","fr","de","gr","hu","ie","it",
            "lv","lt","lu","mt","nl","pl","pt","ro","sk","si","es","se"}
EU_NAMES = {"austria","belgium","bulgaria","croatia","czechia","czech republic","denmark",
            "estonia","finland","france","germany","greece","hungary","ireland","italy",
            "latvia","lithuania","luxembourg","malta","netherlands","poland","portugal",
            "romania","slovakia","slovenia","spain","sweden"}


def _log(v, lo, hi):
    if v <= 0 or lo <= 0 or hi <= lo: return 0.0
    return max(0.0, min(100.0, (math.log10(max(v,1)) - math.log10(lo)) / (math.log10(hi) - math.log10(lo)) * 100))


def _lin(v, lo, hi):
    if hi <= lo: return 0.0
    return max(0.0, min(100.0, (v - lo) / (hi - lo) * 100))


def _cat(cats, targets):
    if not targets: return 0.0
    hcats = [c.lower() for c in (cats or [])]
    m = sum(1 for tc in targets if any(tc in hc or hc in tc for hc in hcats))
    return min(100.0, m / len(targets) * 100)


def score_and_export(
    enricher_checkpoint="", output_path="", categories=None,
    priority_threshold=50, secondary_threshold=25
):
    cats = [c.lower().strip() for c in (categories or ["ai","developer tools","saas","productivity"])]
    if not enricher_checkpoint:
        enricher_checkpoint = str(Path.home() / "Downloads" / "ph_enricher_checkpoint.json")

    with open(enricher_checkpoint, "r", encoding="utf-8") as f:
        cache = json.load(f)
    hunters = list(cache.values())

    # Score
    for h in hunters:
        xc = int(h.get("x_followers", 0) or 0)
        raw = (_log(h.get("ph_followers", 0), 1000, 50000) * 25 +
               _lin(h.get("products_hunted", 0), 10, 100) * 15 +
               _cat(h.get("categories", []), cats) * 25 +
               _log(h.get("avg_upvotes_last_10", 0), 50, 5000) * 20 +
               _log(xc, 100, 100000) * 10)
        score = round(raw / 95)
        h["score"] = score
        country = h.get("country") or ""
        if country: h["is_eu"] = country.lower() in EU_CODES | EU_NAMES
        h["segment"] = ("Priority" if score >= priority_threshold else
                        "Secondary" if score >= secondary_threshold else "Below Threshold")

    # Sort
    SEG = {"Priority":0,"Secondary":1,"Below Threshold":2}
    hunters.sort(key=lambda h: (SEG.get(h.get("segment",""),3), -int(h.get("score",0) or 0)))

    # Excel
    if not output_path:
        output_path = str(Path.home() / "Downloads" / ("ph_hunters_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"))

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "PH Hunters"
    HDRS = ["Status","Name","PH Profile URL","LinkedIn URL","LinkedIn?","Score","Segment",
            "PH Followers","X Followers","Twitter Handle","Categories","Geo",
            "Last Activity","Products Hunted","Avg Upvotes"]
    WIDTHS = [12,22,38,45,10,8,14,14,12,18,40,18,14,16,14]
    hf = Font(bold=True,color="FFFFFF",name="Calibri"); hfill = PatternFill(start_color="1E2A3A",end_color="1E2A3A",fill_type="solid")
    for ci,(h,w) in enumerate(zip(HDRS,WIDTHS),1):
        c=ws.cell(row=1,column=ci,value=h); c.font=hf; c.fill=hfill
        c.alignment=Alignment(horizontal="center",vertical="center"); ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=30
    pri_f=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
    sec_f=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")
    blw_f=PatternFill(start_color="F5F5F5",end_color="F5F5F5",fill_type="solid")
    url_f=Font(color="0563C1",underline="single",name="Calibri")
    for ri,h in enumerate(hunters,2):
        seg=h.get("segment",""); rfill=pri_f if seg=="Priority" else (sec_f if seg=="Secondary" else blw_f)
        status="⭐ Priority" if seg=="Priority" else ("✅ Secondary" if seg=="Secondary" else "—")
        geo=("🇪🇺 "+h.get("country","")) if h.get("is_eu") else (h.get("country") or "—")
        data=[status,h.get("name",h.get("username","")),h.get("ph_url",""),h.get("linkedin_url") or "",
              "✅" if h.get("linkedin_url") else "❌",int(h.get("score",0) or 0),seg or "—",
              int(h.get("ph_followers",0) or 0),int(h.get("x_followers",0) or 0),h.get("twitter_username") or "",
              ", ".join(str(c) for c in (h.get("categories") or [])[:5]),geo,
              str(h.get("last_activity_date") or "")[:10] or "—",
              int(h.get("products_hunted",0) or 0),round(h.get("avg_upvotes_last_10",0) or 0,1)]
        for ci,v in enumerate(data,1):
            c=ws.cell(row=ri,column=ci,value=v); c.fill=rfill; c.alignment=Alignment(vertical="center")
            if ci in (3,4) and v and str(v).startswith("http"): c.hyperlink=str(v); c.font=url_f
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(HDRS))}{len(hunters)+1}"
    wb.save(output_path)
    p=sum(1 for h in hunters if h.get("segment")=="Priority")
    s=sum(1 for h in hunters if h.get("segment")=="Secondary")
    print(f"Exported {len(hunters)} hunters → {output_path}")
    print(f"⭐{p} Priority | ✅{s} Secondary | —{len(hunters)-p-s} Below | max_score={max(h.get('score',0) for h in hunters)}")
    return output_path
